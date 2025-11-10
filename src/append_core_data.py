import openpyxl
from openpyxl.utils import get_column_letter
import os

def read_columns_from_sheet(file_path, sheet_name, columns):
    """
    Reads specific columns from a given sheet in an XLSX file.
    
    Args:
        file_path (str): Path to the source XLSX file.
        sheet_name (str): Name of the sheet to read from.
        columns (list of str): List of column headers to extract.
        
    Returns:
        list: A list of rows containing the values from the specified columns.
    """
    workbook = openpyxl.load_workbook(file_path, data_only=True)
    sheet = workbook[sheet_name]

    # Get the header row
    header = [cell.value for cell in next(sheet.iter_rows(max_row=1))]
    
    # Find the indices of the desired columns
    col_indices = [header.index(col) for col in columns]

    # Extract data for the specified columns
    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        data.append([row[i] for i in col_indices])

    return data

def append_data_to_xlsx(target_file, data):
    """
    Appends data to an existing XLSX file.
    
    Args:
        target_file (str): Path to the target XLSX file.
        data (list): Data to append, as a list of rows.
    """
    if not os.path.exists(target_file):
        # If the target file does not exist, create it.
        workbook = openpyxl.Workbook()
        sheet = workbook.active
    else:
        workbook = openpyxl.load_workbook(target_file)
        sheet = workbook.active

    # Determine the starting column (first empty column)
    start_col = sheet.max_column + 1

    # Write the data starting from row 5
    start_row = 5

    for row_index, row in enumerate(data, start=start_row):
        for col_index, value in enumerate(row, start=start_col):
            sheet.cell(row=row_index, column=col_index, value=value)

    workbook.save(target_file)

if __name__ == "__main__":
    source_file = "./F204RPM.xlsx"
    sheet_name = "100b"
    target_file = "./Book1.xlsx"

    # Columns to extract
    columns_to_extract = [
        "Core 2 VID [V]", 
        "Core 2 T0 Usage [%]", 
        "Core 2 T1 Usage [%]", 
        "Core 2 [°C]",
        "Core 25 VID [V]", 
        "Core 25 T0 Usage [%]", 
        "Core 25 T1 Usage [%]", 
        "Core 25 [°C]",
        "Core 6 VID [V]", 
        "Core 6 T0 Usage [%]", 
        "Core 6 T1 Usage [%]", 
        "Core 6 [°C]"
    ]

    # Read specific columns from the source file
    data = read_columns_from_sheet(source_file, sheet_name, columns_to_extract)

    # Append the data to the target file
    append_data_to_xlsx(target_file, data)

    print(f"Data from {source_file} (sheet: {sheet_name}) has been appended to {target_file}.")
